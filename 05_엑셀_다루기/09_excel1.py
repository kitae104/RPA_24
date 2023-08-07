from openpyxl import Workbook
wb = Workbook()
print(wb.sheetnames)

ws = wb.active

ws2 = wb.create_sheet("new_sheet2");    # 마지막에 시트 추가
ws1 = wb.create_sheet("new_sheet1",1)   # 두번째에 시트 삽입
print(wb.sheetnames)

ws = wb['Sheet']      # 워크시트 선택
ws.title = '주소'
print(ws.title)

ws['A1'] = '이름'
ws['B1'] = '전화번호'

ws['A2'] = '홍길동'
ws['B2'] = '7777'

ws.cell(row=3, column=1, value='홍길순');  # A3
ws.cell(row=3, column=2, value='3333');   # B3

ws.cell(row=3, column=1).value = '홍길순'  # A3
ws.cell(row=3, column=2).value = '3333'   # B3

cell = ws.cell(row=3, column=1)
cell.value = '홍길순'

wb.save('address.xlsx')